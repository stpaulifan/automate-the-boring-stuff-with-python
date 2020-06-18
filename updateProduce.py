#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produkte.xlsx')
sheet = wb.get_active_sheet()


print(sheet)
print(sheet.max_row)

# The produce types and their updated prices
PRICE_UPDATES = {'Glas': 4.27,
                 'Eimer': 2.49,
                 'Affe': 9.99,
                 'Test': 1.164}

# TODO: Loop through the rows and update the prices.

for rowNum in range(2, 7):  # skip the first row
     produceName = sheet.cell(row=rowNum, column=1).value
     if produceName in PRICE_UPDATES:
           sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('produkte.xlsx')